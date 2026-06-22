# 使用 openpyxl 在 Excel 中生成图表
from logginger import get_logger
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, Reference

logger = get_logger(__name__)

SELLER_SHEET = 'seller_summary'
CATEGORY_SHEET = 'category_summary'
CHARTS_SHEET = 'charts'

# 整体画图过程
"""
1. 通过Path方法，判断输出路径是否存在，如果不存在，则创建一个Excel文件
2. 通过load_workbook方法，加载Excel文件
3. 通过sheetnames方法，判断表格中是否存在图表页，如果存在，则删除它，每次做图都需要先判断
4. 通过create_sheet方法，创建一个sheet页来存放图表，这个sheet页的名称是charts。
    如果是需要新的sheet页，那么就需要创建新的。
    如果在已存在的某个sheet页面上创建，那么不需要创建新的，直接在某个页面创建就行了
5. 通过build_seller_pie_chart方法，画饼图
6. 创建一个表，例如饼图 bar_chart = PieChart()
7. 设置图表的标题，例如 chart.title = '销售业绩占比'、labels、data等信息
8. 把数据和分类绑定到图表，例如 chart.add_data(data, titles_from_data=True)、chart.set_categories(labels)
9. 把图表添加到工作表，例如 ws_charts.add_chart(chart, 'A1')
10. 保存文件，例如 wb.save(output_path)
"""

# 画图的时候不需要读取数据，只需要从表中获取sheet页作为数据的来源即可
def add_chart_to_workbook(output_path):
  output_path = Path(output_path)
  if not output_path.exists():
    logger.error('输出路径不存在: %s', output_path)
    raise FileNotFoundError(f'输出路径不存在: {output_path}')

  wb = load_workbook(output_path)

  # 这行代码的意义是：如果表格中存在这个sheet页，则删除它，因为存在重复跑文件的情况，不删除的话，会报错的
  if CHARTS_SHEET in wb.sheetnames:
    logger.info('删除已存在的图表页: %s', CHARTS_SHEET)
    del wb[CHARTS_SHEET]
  # 这行代码的意义是：需要创建一个sheet页来存放图表，这个sheet页的名称是charts
  ws_charts = wb.create_sheet(CHARTS_SHEET)

  build_seller_pie_chart(wb, ws_charts)
  build_category_bar_chart(wb, ws_charts)

  wb.save(output_path)
  logger.info('图表已保存到: %s', output_path)

def build_seller_pie_chart(wb, ws_charts):
  """基于 seller_summary 的 seller / total_amount 画饼图"""
  if SELLER_SHEET not in wb.sheetnames:
    raise ValueError(f'缺少工作表: {SELLER_SHEET}')

  ws_data = wb[SELLER_SHEET]
  if ws_data.max_row < 2:
    logger.warning('%s 无数据，跳过饼图', SELLER_SHEET)
    return

  chart = PieChart()
  chart.title = '销售业绩占比'

  # 第 1 列 seller 作为分类，第 2 列 total_amount 作为数值
  labels = Reference(ws_data, min_col=1, min_row=2, max_row=ws_data.max_row)
  data = Reference(ws_data, min_col=2, min_row=1, max_row=ws_data.max_row)

  # 把数据和分类绑定到图表
  # titles_from_data=True 表示使用数据的第一行作为图表的标题
  chart.add_data(data, titles_from_data=True)
  # set_categories 表示使用数据的第一列作为图表的分类
  chart.set_categories(labels)
  # 把图表添加到工作表
  # A1 表示图表的左上角位置
  ws_charts.add_chart(chart, 'A1')
  logger.info('销售业绩饼图创建完成')

def build_category_bar_chart(wb, ws_charts):
  """基于 category_summary 的 category / total_amount 画柱状图"""
  if CATEGORY_SHEET not in wb.sheetnames:
    raise ValueError(f'缺少工作表: {CATEGORY_SHEET}')

  ws_data = wb[CATEGORY_SHEET]
  if ws_data.max_row < 2:
    logger.warning('%s 无数据，跳过柱状图', CATEGORY_SHEET)
    return

  chart = BarChart()
  chart.type = 'col'
  chart.title = '类目销售额'
  chart.x_axis.title = '类目'
  chart.y_axis.title = '销售额'

  # 第 2 列 total_amount 是数值，第 1 列 category 是横轴分类
  data = Reference(ws_data, min_col=2, min_row=1, max_row=ws_data.max_row)
  categories = Reference(ws_data, min_col=1, min_row=2, max_row=ws_data.max_row)

  chart.add_data(data, titles_from_data=True)
  chart.set_categories(categories)

  ws_charts.add_chart(chart, 'J1')
  logger.info('类目销售额柱状图创建完成')
